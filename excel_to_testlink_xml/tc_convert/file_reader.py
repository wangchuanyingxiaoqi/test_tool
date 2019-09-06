import openpyxl

from .xml_writer import write_xml
from .common import TEST_SUITE, NAME, IMPORTANCE, TEST_TYPE, KEY_WORD, SUMMARY, PRECONDITIONS, STEP_NUMBER, \
    ACTIONS, EXPECTRESULTS, STEPS, STEP, TC_NAME, TC_STEPS


def get_upper_name(ws, cell, num):
    if num == 0:
        return None

    cur_cell = cell
    tc_name = ws[cur_cell.row][cur_cell.column - num].value
    if tc_name is None:
        i = cell.row
        while i >= 0:
            tc_name = ws[i][cur_cell.column - num].value
            if tc_name is not None:
                return tc_name
            i -= 1
    return tc_name


def get_tc_info_dict(ws):
    upper_step = 2
    tc_column = 'B'
    if ws[1][0].value == TC_NAME:
        upper_step = 0
        tc_column = 'A'

    return {
        a.value: {NAME: a.value, IMPORTANCE: ws[a.row][a.column].value, TEST_TYPE: ws[a.row][a.column + 1].value,
                  KEY_WORD: ws[a.row][a.column + 2].value,
                  SUMMARY: ws[a.row][a.column + 3].value, PRECONDITIONS: ws[a.row][a.column + 4].value, STEPS: [],
                  TEST_SUITE: get_upper_name(ws, a, upper_step)
                  }
        for a in ws[tc_column] if
        (a.value is not None and a.value != TC_NAME)}


def get_tc_step_data_dict(ws):
    tc_column = 'H'
    if ws[1][0].value == TC_NAME:
        tc_column = 'G'

    return [
        {get_upper_name(ws, a, 7):
             {STEP_NUMBER: 1, ACTIONS: a.value, EXPECTRESULTS: ws[a.row][a.column].value}} for a in ws[tc_column] if
        (a.value is not None and a.value != TC_STEPS)]


def get_ts_dict(tc_step_data_dict, tc_info_dict):
    ts_dict = {}

    # 整合测试用例
    for ele in tc_step_data_dict:
        n = len(tc_info_dict[list(ele.keys())[0]]['steps'])
        step_info = ele[list(ele.keys())[0]]
        step_info[STEP_NUMBER] = n + 1
        tc_info_dict[list(ele.keys())[0]]['steps'].append({STEP: step_info})

    # 得到testsuite和tc完整映射
    for k, v in tc_info_dict.items():
        tc = v[TEST_SUITE]
        v.pop(TEST_SUITE)
        if ts_dict.get(tc) is None:
            tc_list = []
            tc_list.append(v)
            ts_dict[tc] = tc_list
        else:
            ts_dict[tc].append(v)

    print(ts_dict)
    return ts_dict
